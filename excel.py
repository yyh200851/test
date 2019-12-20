#!C:\python27
import openpyxl
from decimal import *
from openpyxl import Workbook
from openpyxl import styles
from openpyxl.styles import colors
from openpyxl.styles import Font
i=raw_input("please choose the case:【1】数据化整【2】化整查验【3】quit")
i=int(i)
print 'case:',i,"runing"
if i==3:
    exit();
# 打开excel文件,获取工作簿对象
wb = openpyxl.load_workbook('D:/python/excel/sheet1.xlsx')
# 从表单中获取单元格的内容
ws = wb.active  # 当前活跃的表单
ft=Font(color=colors.RED)
#ws['A1'].font=ft
m=1
n=1
sympol=0
E=[1,4,7,10,13,16,19,22,25,28,31,34,37]
print("input the processing interval")
d=raw_input("kn=")
print("input the digit number")
e=int(raw_input("d="))
for m in range(1,39):
    print ('level',m)
    if m in E:
        while (isinstance( ws.cell(row=m,column=n).value,float)|( ws.cell(row=m,column=n).value==0)):
            x=ws.cell(row=m, column=n).value# 获取第m行第n列的单元格
            if x<0:
                sympol=1
            else:
                sympol=0
            x=abs(x)
            #print(x)
            x=float(x);
            d=float(d);
            y=int(x/d)
            #print(y)
            z=round(Decimal(x)-Decimal(d*y),5)
            if z > (0.5*d):
                r=d*(y+1)
                #print(r)
            else:
                if z<(0.5*d):
                    r=d*y
                   #print(r)
                else:
                    if   (y%2==0):
                        r=d*y
                        #print(r)
                    else:
                        r=d*(y+1)
                        #print(r)
            #print(ws.cell(row=(m+2), column=n).value)
            if i==1:
                if sympol==1:
                    ws.cell(row=(m+2),column=n).value=round(-r,4)
                else:
                    ws.cell(row=(m+2),column=n).value=round(r,4)                    
            elif i==2:
                if r!=abs(ws.cell(row=(m+2), column=n).value):
                    ws.cell(row=(m+2),column=n).font=ft
                    #print("adom")
            n=n+1
        else:
            m=m+1;
            n=1
wb.save("D:/python/excel/sheet1.xlsx")
print("finish!")

